# Author : Sviatoslav Besnard pro@slavi.dev
# Created at: 2022-08-09

import sys

import pandas as pd

sys.path.append('../')

from enum import Enum
from typing import Optional
import re

import typer

from utils.utils import RichConsole

import openpyxl as xl

app = typer.Typer(rich_markup_mode="rich")


class RowStopStrategy(str, Enum):
    """
    Strategy to stop the process.
    """
    on_nan = "on_nan"
    end_of_tab = "end_of_tab"
    index_row = "index_row"


class ColumnResultStrategy(str, Enum):
    """
    Strategy to store the result.
    """
    insert_column_right = "insert_column_right"
    add_column_end = "add_column_end"
    index_column = "index_column"


def convert_xlsb_to_xlsx(input_file):
    """
    Convert xlsb to xlsx.
    :param input_file: xlsb file
    :return: xlsx file
    """
    # use pandas to convert xlsb to xlsx
    df = pd.read_excel(input_file, engine='pyxlsb')
    df.to_excel(input_file.replace('.xlsb', '.xlsx'), index=False)


@app.command()
def table_import(
        input_file: str = typer.Argument(..., help="Input file (.xlsx or .xlsb)"),
        output_file: Optional[str] = typer.Option(None, "--output", "-o",
                                                  help="Output file (.xlsx). "
                                                       "If not specified, the input file will be overwritten. "
                                                       "Usage: `-o output.xlsx`"),
        tabs: list[str] = typer.Option([], "--tabs", "-t",
                                       help="Tab name to process, can be specified multiple times.\n"
                                            "Usage: `-t 'sheet 1'`"),
        tab_index: list[str] = typer.Option([], "--tabs_index", "-i",
                                            help="Tab index (position) to process, can be specified multiple times.\n"
                                                 "Can be specified with -t option.\n"
                                                 "Usage: `-i 5 -i 7`"),
        column_pattern: str = typer.Option("^Amount.*", "--column_pattern", "--col_regex",
                                           help="Regex pattern column name to process the amounts. "
                                                "Take the first column that match the pattern. "
                                                "Option ignored if --column_index is specified.\n"
                                                "Usage: `-p '^Amount.*'`"),
        column_index: Optional[int] = typer.Option(None, "--column_index", "--col_idx",
                                                   help="Column index to process (start from 1). "
                                                        "If specified, ignore --column_pattern.\n"
                                                        "Usage: `-i 5`"),
        row_start_index: int = typer.Option(2, "--row_start_index", "--row_start",
                                            help="Row index (start from 1). "
                                                 "Useful if the first row is not the header.\n"
                                                 "Usage: `-s 2`"),
        row_stop_strategy: RowStopStrategy =
        typer.Option(RowStopStrategy.on_nan, "--row_stop_strategy", "--row_stop",
                     help="Strategy to stop the process. "
                          "'on_nan' stops the process when a cell is Not a Number. "
                          "'end_of_tab' stops the process when the last row is reached, ignore NaN values. "
                          "'index_row' stops the process when the row index(included) is reached, ignore NaN values, "
                          "must be used with --row_stop_index.\n"
                          "Usage: `--row_stop on_nan`"),
        row_stop_index: Optional[int] = typer.Option(None, "--row_stop_index",
                                                     help="Row index (included) to stop the process (start from 1). "
                                                          "Must be used with --row_stop_strategy index_row.\n"
                                                          "Usage: `--row_stop_strategy index_row --row_stop_index 42`"),
        column_result_strategy: ColumnResultStrategy =
        typer.Option(ColumnResultStrategy.insert_column_right,
                     "--column_result_strategy", "--col_result",
                     help="Strategy to store the result.\n"
                          "'insert_column_right' inserts the result in a new column to the right of the processed column.\n"
                          "'add_column_end' adds the result in a new column at the end of the sheet.\n"
                          "'index_column' inserts the result in a specific column (will override preexistent data).\n"
                          "Usage: `--col_result insert_column_right`"),
        column_result_index: Optional[int] = typer.Option(None, "--column_result_index", "--col_result_idx",
                                                          help="Column index to store the result (start from 1). "
                                                               "Must be used with --col_result index_column.\n"
                                                               "Usage: `--col_result index_column --col_result_index 42`"),
        partition_column_index: Optional[int] = typer.Option(None, "--partition_column_index", "--part_col_idx",
                                                             help="Column index to partition rows (start from 1). "
                                                                  "If not specified, the whole sheet is processed in on block.\n"
                                                                  "Usage: `--part_col_idx 5`"),
        author: bool = typer.Option(False, "--author", "-a",
                                    help="Print author information and exist.\n"
                                         "Full usage: `python cli.py a -a`"),
        verbose: int = typer.Option(1, "--verbose", "-v",
                                    help="Verbose level (0=error/warning, 1=general, 2=all). "
                                         "Usage: `-v 1`"),
):
    """
        Process comparison between lists of values and pair opposite values.

        Typical usage:

        `python cli.py input.xlsx -o output.xlsx -t 'sheet 1'`

        [blue]Author:
        [cyan]Name: [white]Sviatoslav Besnard
        [cyan]Position: [white]Data Analyst Trainee
        [cyan]Email: [white]pro@slavi.dev

        [blue]Original idea and draft script:
        [cyan]Name: [white]Hugo Grau
        [cyan]Email: [white]hugo.grau@radissonhotels.com
        """
    RichConsole.init()
    RichConsole.verbose = verbose

    # Author
    if author:
        RichConsole.print_author()
        return

    # Checking all the options are valid
    if row_stop_strategy == RowStopStrategy.index_row and row_stop_index is None:
        raise ValueError("--row_stop_index must be specified with --row_stop_strategy index_row")
    if row_stop_strategy != RowStopStrategy.index_row and row_stop_index is not None:
        raise ValueError("--row_stop_index must not be specified with other than --row_stop_strategy index_row")
    if len(tabs) == 0 and len(tab_index) == 0:
        raise ValueError("-t or -i must be specified")
    if column_index is not None and column_index < 1:
        raise ValueError("--column_index must be greater than 0")
    if row_start_index < 1:
        raise ValueError("--row_start_index must be greater than 0")
    if verbose < 0 or verbose > 2:
        raise ValueError("--verbose must be between 0 and 2")

    # if the file is xlsb, convert it to xlsx
    if input_file.endswith(".xlsb"):
        convert_xlsb_to_xlsx(input_file)
        input_file = input_file.replace(".xlsb", ".xlsx")
        RichConsole.info(f"Converted {input_file} to {input_file}")



    # Read the input file
    main_spinner_task = RichConsole.one_shot_task("Opening file")
    wb = None
    try:
        wb = xl.load_workbook(input_file)
    except Exception as e:
        RichConsole.error(f"Error reading input file {input_file}:\n"
                          f"{e}")
        exit(1)
    main_spinner_task.end()

    # Check the input file
    main_spinner_task = RichConsole.one_shot_task("Checking file")
    if len(wb.sheetnames) == 0:
        RichConsole.error("Input file is empty")
        exit(1)

    # Check sheet names
    tab_index = [int(x) for x in tab_index]
    # add the sheet index corresponding name to the list of sheet names
    tabs.extend([name for idx, name in enumerate(wb.sheetnames) if idx + 1 in tab_index])
    # remove duplicates
    tabs = set(tabs)
    sheet_names = set(wb.sheetnames)
    sheets = []
    for tab in tabs:
        if tab not in sheet_names:
            RichConsole.warning(f"Tab '{tab}' not found in input file")
        else:
            sheets.append(wb[tab])
    if len(sheets) == 0:
        RichConsole.error("No matching tab found")
        exit(1)

    # Check column names for each sheet
    for sheet in sheets:
        # Get the column names
        col_names = [cell.value for cell in sheet[1]]
        # Check the column names
        if column_index is not None:
            if column_index > len(col_names):
                RichConsole.error(f"Column index {column_index} not found in sheet '{sheet.title}'")
                exit(1)
            column_name = col_names[column_index - 1]
        else:
            column_name = next((name for name in col_names if re.match(column_pattern, name)), None)
        if column_name is None:
            RichConsole.error(f"Column name not found in sheet '{sheet.title}'")
            exit(1)
        RichConsole.debug(f"Found column '{column_name}' in sheet '{sheet.title}'")

    main_spinner_task.end()

    # Process each sheet
    main_bar_task = RichConsole.progress_bar.add_task("Starting processing", visible=True, total=len(sheets))
    for sheet in sheets:
        RichConsole.debug(f"Processing sheet '{sheet.title}'")
        RichConsole.progress_bar.update(main_bar_task, description=f"{sheet.title}")

        # Get the column names
        col_names = [cell.value for cell in sheet[1]]
        if column_index is not None:
            column_name = col_names[column_index - 1]
        else:
            column_name = next((name for name in col_names if re.match(column_pattern, name)), None)
        # Get the column index
        column_index_process = col_names.index(column_name) + 1

        # get the row data and store it in hash map
        # {abs_cell_value: {"+": list of positive cells, "-": list of negative cells}, ...}
        hash_map = {}
        for cell in sheet.iter_rows(min_row=row_start_index, max_row=sheet.max_row, min_col=column_index_process,
                                    max_col=column_index_process):
            cell = cell[0]
            if row_stop_strategy == RowStopStrategy.index_row and cell.row > row_stop_index:
                break
            if cell.value is None or cell.value == "":
                if row_stop_strategy == RowStopStrategy.on_nan:
                    break
                elif row_stop_strategy == RowStopStrategy.end_of_tab:
                    continue

            partition_value = ""
            if partition_column_index is not None:
                partition_value = sheet.cell(row=cell.row, column=partition_column_index).value

            v = str(abs(cell.value)) + partition_value
            if v not in hash_map:
                hash_map[v] = {"+": [], "-": []}
            if cell.value < 0:
                hash_map[v]["-"].append(cell)
            else:
                hash_map[v]["+"].append(cell)

        # Result column
        index_column_result = None
        if column_result_strategy == ColumnResultStrategy.insert_column_right:
            index_column_result = column_index_process + 1
            sheet.insert_cols(index_column_result)
        elif column_result_strategy == ColumnResultStrategy.add_column_end:
            index_column_result = sheet.max_col + 1
            sheet.insert_cols(index_column_result)
        elif column_result_strategy == ColumnResultStrategy.index_column:
            index_column_result = column_result_index

        for idx, hash_map_value in enumerate(hash_map.items()):
            abs_amount, cells = hash_map_value
            for index_identical_cell in range(max(len(cells["+"]), len(cells["-"]))):
                if len(cells["+"]) > index_identical_cell and len(cells["-"]) > index_identical_cell:
                    cell_plus = cells["+"][index_identical_cell]
                    cell_minus = cells["-"][index_identical_cell]

                    cell_plus_result = sheet.cell(row=cell_plus.row, column=index_column_result)
                    cell_minus_result = sheet.cell(row=cell_minus.row, column=index_column_result)

                    cell_plus_result.value = idx
                    cell_minus_result.value = idx
                elif len(cells["+"]) > index_identical_cell:
                    # leave the result cell empty if there is no cell for the minus
                    cell_plus = cells["+"][index_identical_cell]
                    cell_plus_result = sheet.cell(row=cell_plus.row, column=index_column_result)
                    cell_plus_result.value = None
                elif len(cells["-"]) > index_identical_cell:
                    # leave the result cell empty if there is no cell for the plus
                    cell_minus = cells["-"][index_identical_cell]
                    cell_minus_result = sheet.cell(row=cell_minus.row, column=index_column_result)
                    cell_minus_result.value = None

        RichConsole.progress_bar.update(main_bar_task, advance=1)
    RichConsole.progress_bar.remove_task(main_bar_task)
    RichConsole.one_shot_task("Processing main job").end()

    main_spinner_task = RichConsole.one_shot_task("Saving file")

    # Save the result
    if output_file is None:
        output_file = input_file
    wb.save(output_file)

    main_spinner_task.end()
    RichConsole.live.refresh()


if __name__ == "__main__":
    app()
